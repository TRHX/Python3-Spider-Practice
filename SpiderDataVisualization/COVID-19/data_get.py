# =============================================
# --*-- coding: utf-8 --*--
# @Time    : 2020-07-06
# @Author  : TRHX
# @Blog    : www.itrhx.com
# @CSDN    : itrhx.blog.csdn.net
# @FileName: data_get.py
# @Software: PyCharm
# =============================================

import requests
import json
import openpyxl
from lxml import etree


def init():
    headers = {
        'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/84.0.4147.13 Safari/537.36'
    }
    url = 'https://voice.baidu.com/act/newpneumonia/newpneumonia/'
    response = requests.get(url=url, headers=headers)
    tree = etree.HTML(response.text)
    dict1 = tree.xpath('//script[@id="captain-config"]/text()')
    dict2 = json.loads(dict1[0])
    return dict2


def china_total_data(data):

    """
    1、中国省/直辖市/自治区/行政区疫情数据
    省/直辖市/自治区/行政区：area
    现有确诊：    curConfirm
    累计确诊：    confirmed
    累计治愈：    crued
    累计死亡：    died
    现有确诊增量： curConfirmRelative
    累计确诊增量： confirmedRelative
    累计治愈增量： curedRelative
    累计死亡增量： diedRelative
    """

    wb = openpyxl.Workbook()              # 创建工作簿
    ws_china = wb.active                  # 获取工作表
    ws_china.title = "中国省份疫情数据"   # 命名工作表
    ws_china.append(['省/直辖市/自治区/行政区', '现有确诊', '累计确诊', '累计治愈',
                     '累计死亡', '现有确诊增量', '累计确诊增量',
                     '累计治愈增量', '累计死亡增量'])
    china = data['component'][0]['caseList']
    for province in china:
        ws_china.append([province['area'],
                        province['curConfirm'],
                        province['confirmed'],
                        province['crued'],
                        province['died'],
                        province['curConfirmRelative'],
                        province['confirmedRelative'],
                        province['curedRelative'],
                        province['diedRelative']])

    """
    2、中国城市疫情数据
    城市：city
    现有确诊：curConfirm
    累计确诊：confirmed
    累计治愈：crued
    累计死亡：died
    累计确诊增量：confirmedRelative
    """

    ws_city = wb.create_sheet('中国城市疫情数据')
    ws_city.append(['城市', '现有确诊', '累计确诊',
                    '累计治愈', '累计死亡', '累计确诊增量'])
    for province in china:
        for city in province['subList']:
            # 某些城市没有 curConfirm 数据，则将其设置为 0，crued 和 died 为空时，替换成 0
            if 'curConfirm' not in city:
                city['curConfirm'] = '0'
            if city['crued'] == '':
                city['crued'] = '0'
            if city['died'] == '':
                city['died'] = '0'
            ws_city.append([city['city'], '0', city['confirmed'],
                           city['crued'], city['died'], city['confirmedRelative']])

    """
    3、中国疫情数据更新时间：mapLastUpdatedTime
    """

    time_domestic = data['component'][0]['mapLastUpdatedTime']
    ws_time = wb.create_sheet('中国疫情数据更新时间')
    ws_time.column_dimensions['A'].width = 22  # 调整列宽
    ws_time.append(['中国疫情数据更新时间'])
    ws_time.append([time_domestic])

    wb.save('COVID-19-China.xlsx')
    print('中国疫情数据已保存至 COVID-19-China.xlsx！')


def global_total_data(data):

    """
    1、全球各国疫情数据
    国家：country
    现有确诊：curConfirm
    累计确诊：confirmed
    累计治愈：crued
    累计死亡：died
    累计确诊增量：confirmedRelative
    """

    wb = openpyxl.Workbook()
    ws_global = wb.active
    ws_global.title = "全球各国疫情数据"

    # 按照国家保存数据
    countries = data['component'][0]['caseOutsideList']
    ws_global.append(['国家', '现有确诊', '累计确诊', '累计治愈', '累计死亡', '累计确诊增量'])
    for country in countries:
        ws_global.append([country['area'],
                          country['curConfirm'],
                          country['confirmed'],
                          country['crued'],
                          country['died'],
                          country['confirmedRelative']])

    # 按照洲保存数据
    continent = data['component'][0]['globalList']
    for area in continent:
        ws_foreign = wb.create_sheet(area['area'] + '疫情数据')
        ws_foreign.append(['国家', '现有确诊', '累计确诊', '累计治愈', '累计死亡', '累计确诊增量'])
        for country in area['subList']:
            ws_foreign.append([country['country'],
                               country['curConfirm'],
                               country['confirmed'],
                               country['crued'],
                               country['died'],
                               country['confirmedRelative']])

    # 在“全球各国疫情数据”和“亚洲疫情数据”两张表中写入中国疫情数据
    ws1, ws2 = wb['全球各国疫情数据'], wb['亚洲疫情数据']
    original_data = data['component'][0]['summaryDataIn']
    add_china_data = ['中国',
                      original_data['curConfirm'],
                      original_data['confirmed'],
                      original_data['cured'],
                      original_data['died'],
                      original_data['confirmedRelative']]
    ws1.append(add_china_data)
    ws2.append(add_china_data)

    """
    2、全球疫情数据更新时间：foreignLastUpdatedTime
    """

    time_foreign = data['component'][0]['foreignLastUpdatedTime']
    ws_time = wb.create_sheet('全球疫情数据更新时间')
    ws_time.column_dimensions['A'].width = 22  # 调整列宽
    ws_time.append(['全球疫情数据更新时间'])
    ws_time.append([time_foreign])

    wb.save('COVID-19-Global.xlsx')
    print('全球疫情数据已保存至 COVID-19-Global.xlsx！')


def china_daily_data(data):

    """
    i_dict = data['component'][0]['trend']
    i_dict['updateDate']：日期
    i_dict['list'][0]：确诊
    i_dict['list'][1]：疑似
    i_dict['list'][2]：治愈
    i_dict['list'][3]：死亡
    i_dict['list'][4]：新增确诊
    i_dict['list'][5]：新增疑似
    i_dict['list'][6]：新增治愈
    i_dict['list'][7]：新增死亡
    i_dict['list'][8]：累计境外输入
    i_dict['list'][9]：新增境外输入
    """

    ccd_dict = data['component'][0]['trend']
    update_date = ccd_dict['updateDate']              # 日期
    china_confirmed = ccd_dict['list'][0]['data']     # 每日累计确诊数据
    china_crued = ccd_dict['list'][2]['data']         # 每日累计治愈数据
    china_died = ccd_dict['list'][3]['data']          # 每日累计死亡数据
    wb = openpyxl.load_workbook('COVID-19-China.xlsx')

    # 写入每日累计确诊数据
    ws_china_confirmed = wb.create_sheet('中国每日累计确诊数据')
    ws_china_confirmed.append(['日期', '数据'])
    for data in zip(update_date, china_confirmed):
        ws_china_confirmed.append(data)

    # 写入每日累计治愈数据
    ws_china_crued = wb.create_sheet('中国每日累计治愈数据')
    ws_china_crued.append(['日期', '数据'])
    for data in zip(update_date, china_crued):
        ws_china_crued.append(data)

    # 写入每日累计死亡数据
    ws_china_died = wb.create_sheet('中国每日累计死亡数据')
    ws_china_died.append(['日期', '数据'])
    for data in zip(update_date, china_died):
        ws_china_died.append(data)

    wb.save('COVID-19-China.xlsx')
    print('中国每日累计确诊/治愈/死亡数据已保存至 COVID-19-China.xlsx！')


def foreign_daily_data(data):

    """
    te_dict = data['component'][0]['allForeignTrend']
    te_dict['updateDate']：日期
    te_dict['list'][0]：累计确诊
    te_dict['list'][1]：治愈
    te_dict['list'][2]：死亡
    te_dict['list'][3]：现有确诊
    te_dict['list'][4]：新增确诊
    """

    te_dict = data['component'][0]['allForeignTrend']
    update_date = te_dict['updateDate']                # 日期
    foreign_confirmed = te_dict['list'][0]['data']     # 每日累计确诊数据
    foreign_crued = te_dict['list'][1]['data']         # 每日累计治愈数据
    foreign_died = te_dict['list'][2]['data']          # 每日累计死亡数据
    wb = openpyxl.load_workbook('COVID-19-Global.xlsx')

    # 写入每日累计确诊数据
    ws_foreign_confirmed = wb.create_sheet('境外每日累计确诊数据')
    ws_foreign_confirmed.append(['日期', '数据'])
    for data in zip(update_date, foreign_confirmed):
        ws_foreign_confirmed.append(data)

    # 写入累计治愈数据
    ws_foreign_crued = wb.create_sheet('境外每日累计治愈数据')
    ws_foreign_crued.append(['日期', '数据'])
    for data in zip(update_date, foreign_crued):
        ws_foreign_crued.append(data)

    # 写入累计死亡数据
    ws_foreign_died = wb.create_sheet('境外每日累计死亡数据')
    ws_foreign_died.append(['日期', '数据'])
    for data in zip(update_date, foreign_died):
        ws_foreign_died.append(data)

    wb.save('COVID-19-Global.xlsx')
    print('境外每日累计确诊/治愈/死亡数据已保存至 COVID-19-Global.xlsx！')


if __name__ == '__main__':

    """
    china_total_data:  中国总数据
    global_total_data: 全球总数据
    china_daily_data:  中国每日数据
    foreign_daily_data: 境外每日数据
    注意：全球包含中国，境外不包含中国！
    """

    data_dict = init()
    china_total_data(data_dict)
    global_total_data(data_dict)
    china_daily_data(data_dict)
    foreign_daily_data(data_dict)
