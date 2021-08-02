# =============================================
# --*-- coding: utf-8 --*--
# @Time    : 2020-07-06
# @Author  : TRHX
# @Blog    : www.itrhx.com
# @CSDN    : itrhx.blog.csdn.net
# @FileName: data_map.py
# @Software: PyCharm
# =============================================

from pyecharts.charts import Map, Page, Line
import pyecharts.options as opts
import openpyxl


def china_total_map():
    wb = openpyxl.load_workbook('COVID-19-China.xlsx')  # 获取已有的xlsx文件
    ws_time = wb['中国疫情数据更新时间']                # 获取文件中中国疫情数据更新时间表
    ws_data = wb['中国省份疫情数据']                    # 获取文件中中国省份疫情数据表
    ws_data.delete_rows(1)                              # 删除第一行
    province = []                                       # 省份
    curconfirm = []                                     # 累计确诊
    for data in ws_data.values:
        province.append(data[0])
        curconfirm.append(data[2])
    time_china = ws_time['A2'].value                    # 更新时间

    # 设置分级颜色
    pieces = [
        {'max': 0, 'min': 0, 'label': '0', 'color': '#FFFFFF'},
        {'max': 9, 'min': 1, 'label': '1-9', 'color': '#FFE5DB'},
        {'max': 99, 'min': 10, 'label': '10-99', 'color': '#FF9985'},
        {'max': 999, 'min': 100, 'label': '100-999', 'color': '#F57567'},
        {'max': 9999, 'min': 1000, 'label': '1000-9999', 'color': '#E64546'},
        {'max': 99999, 'min': 10000, 'label': '≧10000', 'color': '#B80909'}
    ]

    # 绘制地图
    ct_map = (
        Map()
        .add(series_name='累计确诊人数', data_pair=[list(z) for z in zip(province, curconfirm)], maptype="china")
        .set_global_opts(
            title_opts=opts.TitleOpts(title="中国疫情数据（累计确诊）",
                                      subtitle='数据更新至：' + time_china + '\n\n来源：百度疫情实时大数据报告'),
            visualmap_opts=opts.VisualMapOpts(max_=300, is_piecewise=True, pieces=pieces)
        )
    )
    return ct_map


def global_total_map():
    wb = openpyxl.load_workbook('COVID-19-Global.xlsx')
    ws_time = wb['全球疫情数据更新时间']
    ws_data = wb['全球各国疫情数据']
    ws_data.delete_rows(1)
    country = []                        # 国家
    curconfirm = []                     # 累计确诊
    for data in ws_data.values:
        country.append(data[0])
        curconfirm.append(data[2])
    time_global = ws_time['A2'].value   # 更新时间

    # 国家名称中英文映射表
    name_map = {
          "Somalia": "索马里",
          "Liechtenstein": "列支敦士登",
          "Morocco": "摩洛哥",
          "W. Sahara": "西撒哈拉",
          "Serbia": "塞尔维亚",
          "Afghanistan": "阿富汗",
          "Angola": "安哥拉",
          "Albania": "阿尔巴尼亚",
          "Andorra": "安道尔共和国",
          "United Arab Emirates": "阿拉伯联合酋长国",
          "Argentina": "阿根廷",
          "Armenia": "亚美尼亚",
          "Australia": "澳大利亚",
          "Austria": "奥地利",
          "Azerbaijan": "阿塞拜疆",
          "Burundi": "布隆迪",
          "Belgium": "比利时",
          "Benin": "贝宁",
          "Burkina Faso": "布基纳法索",
          "Bangladesh": "孟加拉国",
          "Bulgaria": "保加利亚",
          "Bahrain": "巴林",
          "Bahamas": "巴哈马",
          "Bosnia and Herz.": "波斯尼亚和黑塞哥维那",
          "Belarus": "白俄罗斯",
          "Belize": "伯利兹",
          "Bermuda": "百慕大",
          "Bolivia": "玻利维亚",
          "Brazil": "巴西",
          "Barbados": "巴巴多斯",
          "Brunei": "文莱",
          "Bhutan": "不丹",
          "Botswana": "博茨瓦纳",
          "Central African Rep.": "中非共和国",
          "Canada": "加拿大",
          "Switzerland": "瑞士",
          "Chile": "智利",
          "China": "中国",
          "Côte d'Ivoire": "科特迪瓦",
          "Cameroon": "喀麦隆",
          "Dem. Rep. Congo": "刚果（布）",
          "Congo": "刚果（金）",
          "Colombia": "哥伦比亚",
          "Cape Verde": "佛得角",
          "Costa Rica": "哥斯达黎加",
          "Cuba": "古巴",
          "N. Cyprus": "北塞浦路斯",
          "Cyprus": "塞浦路斯",
          "Czech Rep.": "捷克",
          "Germany": "德国",
          "Djibouti": "吉布提",
          "Denmark": "丹麦",
          "Dominican Rep.": "多米尼加",
          "Algeria": "阿尔及利亚",
          "Ecuador": "厄瓜多尔",
          "Egypt": "埃及",
          "Eritrea": "厄立特里亚",
          "Spain": "西班牙",
          "Estonia": "爱沙尼亚",
          "Ethiopia": "埃塞俄比亚",
          "Finland": "芬兰",
          "Fiji": "斐济",
          "France": "法国",
          "Gabon": "加蓬",
          "United Kingdom": "英国",
          "Georgia": "格鲁吉亚",
          "Ghana": "加纳",
          "Guinea": "几内亚",
          "Gambia": "冈比亚",
          "Guinea-Bissau": "几内亚比绍",
          "Eq. Guinea": "赤道几内亚",
          "Greece": "希腊",
          "Grenada": "格林纳达",
          "Greenland": "格陵兰岛",
          "Guatemala": "危地马拉",
          "Guam": "关岛",
          "Guyana": "圭亚那合作共和国",
          "Honduras": "洪都拉斯",
          "Croatia": "克罗地亚",
          "Haiti": "海地",
          "Hungary": "匈牙利",
          "Indonesia": "印度尼西亚",
          "India": "印度",
          "Br. Indian Ocean Ter.": "英属印度洋领土",
          "Ireland": "爱尔兰",
          "Iran": "伊朗",
          "Iraq": "伊拉克",
          "Iceland": "冰岛",
          "Israel": "以色列",
          "Italy": "意大利",
          "Jamaica": "牙买加",
          "Jordan": "约旦",
          "Japan": "日本",
          "Siachen Glacier": "锡亚琴冰川",
          "Kazakhstan": "哈萨克斯坦",
          "Kenya": "肯尼亚",
          "Kyrgyzstan": "吉尔吉斯斯坦",
          "Cambodia": "柬埔寨",
          "Korea": "韩国",
          "Kuwait": "科威特",
          "Lao PDR": "老挝",
          "Lebanon": "黎巴嫩",
          "Liberia": "利比里亚",
          "Libya": "利比亚",
          "Sri Lanka": "斯里兰卡",
          "Lesotho": "莱索托",
          "Lithuania": "立陶宛",
          "Luxembourg": "卢森堡",
          "Latvia": "拉脱维亚",
          "Moldova": "摩尔多瓦",
          "Madagascar": "马达加斯加",
          "Mexico": "墨西哥",
          "Macedonia": "马其顿",
          "Mali": "马里",
          "Malta": "马耳他",
          "Myanmar": "缅甸",
          "Montenegro": "黑山",
          "Mongolia": "蒙古国",
          "Mozambique": "莫桑比克",
          "Mauritania": "毛里塔尼亚",
          "Mauritius": "毛里求斯",
          "Malawi": "马拉维",
          "Malaysia": "马来西亚",
          "Namibia": "纳米比亚",
          "New Caledonia": "新喀里多尼亚",
          "Niger": "尼日尔",
          "Nigeria": "尼日利亚",
          "Nicaragua": "尼加拉瓜",
          "Netherlands": "荷兰",
          "Norway": "挪威",
          "Nepal": "尼泊尔",
          "New Zealand": "新西兰",
          "Oman": "阿曼",
          "Pakistan": "巴基斯坦",
          "Panama": "巴拿马",
          "Peru": "秘鲁",
          "Philippines": "菲律宾",
          "Papua New Guinea": "巴布亚新几内亚",
          "Poland": "波兰",
          "Puerto Rico": "波多黎各",
          "Dem. Rep. Korea": "朝鲜",
          "Portugal": "葡萄牙",
          "Paraguay": "巴拉圭",
          "Palestine": "巴勒斯坦",
          "Qatar": "卡塔尔",
          "Romania": "罗马尼亚",
          "Russia": "俄罗斯",
          "Rwanda": "卢旺达",
          "Saudi Arabia": "沙特阿拉伯",
          "Sudan": "苏丹",
          "S. Sudan": "南苏丹",
          "Senegal": "塞内加尔",
          "Singapore": "新加坡",
          "Solomon Is.": "所罗门群岛",
          "Sierra Leone": "塞拉利昂",
          "El Salvador": "萨尔瓦多",
          "Suriname": "苏里南",
          "Slovakia": "斯洛伐克",
          "Slovenia": "斯洛文尼亚",
          "Sweden": "瑞典",
          "Swaziland": "斯威士兰",
          "Seychelles": "塞舌尔",
          "Syria": "叙利亚",
          "Chad": "乍得",
          "Togo": "多哥",
          "Thailand": "泰国",
          "Tajikistan": "塔吉克斯坦",
          "Turkmenistan": "土库曼斯坦",
          "Timor-Leste": "东帝汶",
          "Tonga": "汤加",
          "Trinidad and Tobago": "特立尼达和多巴哥",
          "Tunisia": "突尼斯",
          "Turkey": "土耳其",
          "Tanzania": "坦桑尼亚",
          "Uganda": "乌干达",
          "Ukraine": "乌克兰",
          "Uruguay": "乌拉圭",
          "United States": "美国",
          "Uzbekistan": "乌兹别克斯坦",
          "Venezuela": "委内瑞拉",
          "Vietnam": "越南",
          "Vanuatu": "瓦努阿图",
          "Yemen": "也门",
          "South Africa": "南非",
          "Zambia": "赞比亚",
          "Zimbabwe": "津巴布韦",
          "Aland": "奥兰群岛",
          "American Samoa": "美属萨摩亚",
          "Fr. S. Antarctic Lands": "南极洲",
          "Antigua and Barb.": "安提瓜和巴布达",
          "Comoros": "科摩罗",
          "Curaçao": "库拉索岛",
          "Cayman Is.": "开曼群岛",
          "Dominica": "多米尼加",
          "Falkland Is.": "福克兰群岛马尔维纳斯",
          "Faeroe Is.": "法罗群岛",
          "Micronesia": "密克罗尼西亚",
          "Heard I. and McDonald Is.": "赫德岛和麦克唐纳群岛",
          "Isle of Man": "曼岛",
          "Jersey": "泽西岛",
          "Kiribati": "基里巴斯",
          "Saint Lucia": "圣卢西亚",
          "N. Mariana Is.": "北马里亚纳群岛",
          "Montserrat": "蒙特塞拉特",
          "Niue": "纽埃",
          "Palau": "帕劳",
          "Fr. Polynesia": "法属波利尼西亚",
          "S. Geo. and S. Sandw. Is.": "南乔治亚岛和南桑威奇群岛",
          "Saint Helena": "圣赫勒拿",
          "St. Pierre and Miquelon": "圣皮埃尔和密克隆群岛",
          "São Tomé and Principe": "圣多美和普林西比",
          "Turks and Caicos Is.": "特克斯和凯科斯群岛",
          "St. Vin. and Gren.": "圣文森特和格林纳丁斯",
          "U.S. Virgin Is.": "美属维尔京群岛",
          "Samoa": "萨摩亚"
        }

    pieces = [
        {'max': 0, 'min': 0, 'label': '0', 'color': '#FFFFFF'},
        {'max': 49, 'min': 1, 'label': '1-49', 'color': '#FFE5DB'},
        {'max': 99, 'min': 50, 'label': '50-99', 'color': '#FFC4B3'},
        {'max': 999, 'min': 100, 'label': '100-999', 'color': '#FF9985'},
        {'max': 9999, 'min': 1000, 'label': '1000-9999', 'color': '#F57567'},
        {'max': 99999, 'min': 10000, 'label': '10000-99999', 'color': '#E64546'},
        {'max': 999999, 'min': 100000, 'label': '100000-999999', 'color': '#B80909'},
        {'max': 9999999, 'min': 1000000, 'label': '≧1000000', 'color': '#8A0808'}
    ]

    gt_map = (
        Map()
        .add(series_name='累计确诊人数', data_pair=[list(z) for z in zip(country, curconfirm)], maptype="world", name_map=name_map, is_map_symbol_show=False)
        .set_series_opts(label_opts=opts.LabelOpts(is_show=False))
        .set_global_opts(
            title_opts=opts.TitleOpts(title="全球疫情数据（累计确诊）",
                                      subtitle='数据更新至：' + time_global + '\n\n来源：百度疫情实时大数据报告'),
            visualmap_opts=opts.VisualMapOpts(max_=300, is_piecewise=True, pieces=pieces),
        )
    )
    return gt_map


def china_daily_map():
    wb = openpyxl.load_workbook('COVID-19-China.xlsx')
    ws_china_confirmed = wb['中国每日累计确诊数据']
    ws_china_crued = wb['中国每日累计治愈数据']
    ws_china_died = wb['中国每日累计死亡数据']

    ws_china_confirmed.delete_rows(1)
    ws_china_crued.delete_rows(1)
    ws_china_died.delete_rows(1)

    x_date = []               # 日期
    y_china_confirmed = []    # 每日累计确诊
    y_china_crued = []        # 每日累计治愈
    y_china_died = []         # 每日累计死亡

    for china_confirmed in ws_china_confirmed.values:
        y_china_confirmed.append(china_confirmed[1])
    for china_crued in ws_china_crued.values:
        x_date.append(china_crued[0])
        y_china_crued.append(china_crued[1])
    for china_died in ws_china_died.values:
        y_china_died.append(china_died[1])

    fi_map = (
        Line(init_opts=opts.InitOpts(height='420px'))
            .add_xaxis(xaxis_data=x_date)
            .add_yaxis(
            series_name="中国累计确诊数据",
            y_axis=y_china_confirmed,
            label_opts=opts.LabelOpts(is_show=False),
        )
            .add_yaxis(
            series_name="中国累计治愈趋势",
            y_axis=y_china_crued,
            label_opts=opts.LabelOpts(is_show=False),
        )
            .add_yaxis(
            series_name="中国累计死亡趋势",
            y_axis=y_china_died,
            label_opts=opts.LabelOpts(is_show=False),
        )
            .set_global_opts(
            title_opts=opts.TitleOpts(title="中国每日累计确诊/治愈/死亡趋势"),
            legend_opts=opts.LegendOpts(pos_bottom="bottom", orient='horizontal'),
            tooltip_opts=opts.TooltipOpts(trigger="axis"),
            yaxis_opts=opts.AxisOpts(
                type_="value",
                axistick_opts=opts.AxisTickOpts(is_show=True),
                splitline_opts=opts.SplitLineOpts(is_show=True),
            ),
            xaxis_opts=opts.AxisOpts(type_="category", boundary_gap=False),
        )
    )
    return fi_map


def foreign_daily_map():
    wb = openpyxl.load_workbook('COVID-19-Global.xlsx')
    ws_foreign_confirmed = wb['境外每日累计确诊数据']
    ws_foreign_crued = wb['境外每日累计治愈数据']
    ws_foreign_died = wb['境外每日累计死亡数据']

    ws_foreign_confirmed.delete_rows(1)
    ws_foreign_crued.delete_rows(1)
    ws_foreign_died.delete_rows(1)

    x_date = []                # 日期
    y_foreign_confirmed = []   # 累计确诊
    y_foreign_crued = []       # 累计治愈
    y_foreign_died = []        # 累计死亡

    for foreign_confirmed in ws_foreign_confirmed.values:
        y_foreign_confirmed.append(foreign_confirmed[1])
    for foreign_crued in ws_foreign_crued.values:
        x_date.append(foreign_crued[0])
        y_foreign_crued.append(foreign_crued[1])
    for foreign_died in ws_foreign_died.values:
        y_foreign_died.append(foreign_died[1])

    fte_map = (
        Line(init_opts=opts.InitOpts(height='420px'))
            .add_xaxis(xaxis_data=x_date)
            .add_yaxis(
            series_name="境外累计确诊趋势",
            y_axis=y_foreign_confirmed,
            label_opts=opts.LabelOpts(is_show=False),
        )
            .add_yaxis(
            series_name="境外累计治愈趋势",
            y_axis=y_foreign_crued,
            label_opts=opts.LabelOpts(is_show=False),
        )
            .add_yaxis(
            series_name="境外累计死亡趋势",
            y_axis=y_foreign_died,
            label_opts=opts.LabelOpts(is_show=False),
        )
            .set_global_opts(
            title_opts=opts.TitleOpts(title="境外每日累计确诊/治愈/死亡趋势"),
            legend_opts=opts.LegendOpts(pos_bottom="bottom", orient='horizontal'),
            tooltip_opts=opts.TooltipOpts(trigger="axis"),
            yaxis_opts=opts.AxisOpts(
                type_="value",
                axistick_opts=opts.AxisTickOpts(is_show=True),
                splitline_opts=opts.SplitLineOpts(is_show=True),
            ),
            xaxis_opts=opts.AxisOpts(type_="category", boundary_gap=False),
        )
    )
    return fte_map


def all_map():

    """
    china_total_map:   中国累计确诊地图
    global_total_map:  全球累计确诊地图
    china_daily_map:   中国每日数据折线图
    foreign_daily_map: 境外每日数据折线图
    全球包含中国，境外不包含中国！
    """

    page = Page(layout=Page.SimplePageLayout, page_title="TRHX丨疫情实时大数据")
    page.add(
        china_total_map(),
        global_total_map(),
        china_daily_map(),
        foreign_daily_map()
    )
    page.render('COVID-19.html')
    print('国内以及境外疫情可视化图表绘制完毕！')


if __name__ == '__main__':
    all_map()
